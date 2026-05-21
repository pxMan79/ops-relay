import subprocess
import json
import pandas as pd
import datetime
import os
import shutil
import socket
import struct
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ================= 配置区 =================
EXCEL_PATH = f"/root/ops_monitor/Server_Report_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
# ==========================================

def run_ansible_tree(group, module, args):
    out_dir = f"/tmp/ansible_out_{group}"
    if os.path.exists(out_dir): shutil.rmtree(out_dir)
    os.makedirs(out_dir)
    cmd = ["ansible", group, "-m", module, "-a", args, "-t", out_dir]
    subprocess.run(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    results = {}
    for fname in os.listdir(out_dir):
        try:
            with open(os.path.join(out_dir, fname), 'r', encoding='utf-8') as f:
                results[fname] = json.load(f)
        except Exception:
            pass
    return results

def format_uptime(secs_str):
    try:
        secs = float(secs_str.split()[0])
        d = int(secs // 86400)
        h = int((secs % 86400) // 3600)
        return f"{d:>2d}天 {h:>2d}小时"
    except:
        return "获取失败"

def clean_os_name(raw_name):
    clean = re.sub(r' Linux release | release ', ' ', raw_name).split(' (')[0].strip()
    match = re.match(r'^([a-zA-Z\s]+ \d+\.\d+)', clean)
    if match: return match.group(1)
    return clean

def get_linux_data():
    cpu_cmd = "(awk '/^cpu / {print $2+$3+$4, $2+$3+$4+$5+$6+$7+$8}' /proc/stat; sleep 0.2; awk '/^cpu / {print $2+$3+$4, $2+$3+$4+$5+$6+$7+$8}' /proc/stat) | awk '{if(NR==1){u1=$1;t1=$2} else if(NR==2){if($2-t1==0) printf \"0.0\\n\"; else printf \"%.1f\\n\", ($1-u1)/($2-t1)*100}}'"
    cmd = f'echo "---OS---"; cat /etc/redhat-release 2>/dev/null; echo "---CPU---"; {cpu_cmd}; echo "---MEM---"; free -m; echo "---DISK---"; df -hP; echo "---UPTIME---"; cat /proc/uptime'
    
    json_data = run_ansible_tree("all_linux", "raw", cmd)
    data_list = []
    for host, info in json_data.items():
        if info.get('failed') or info.get('unreachable'): continue
        stdout = info.get('stdout', '')
        os_ver = cpu = mem_str = sys_disk = data_disk_str = uptime = "获取失败"
        data_disks = []
        mode = ""
        for line in stdout.split('\n'):
            line = line.strip()
            if not line: continue
            if line.startswith("---"): mode = line.replace("-", ""); continue

            if mode == "OS": os_ver = clean_os_name(line); mode = ""
            elif mode == "CPU":
                try: cpu = f"{float(line):>5.1f}%"
                except: cpu = line
                mode = "" 
            elif mode == "MEM" and line.startswith("Mem:"):
                parts = line.split()
                total, used = float(parts[1]), float(parts[2])
                remain = total - used
                mem_pct = (used/total)*100 if total > 0 else 0
                mem_str = f"{mem_pct:>4.1f}% (余 {remain/1024:>5.1f}G / 共 {total/1024:>5.1f}G)"
            elif mode == "DISK" and line.startswith("/dev/"):
                parts = line.split()
                size, used, avail, use_pct, mount = parts[1], parts[2], parts[3], parts[4], parts[5]
                if mount == "/":
                    sys_disk = f"{use_pct:>4} (余 {avail:>5} / 共 {size:>5})"
                elif not mount.startswith("/boot"):
                    data_disks.append(f"{mount+':':<12} {use_pct:>4} (余 {avail:>5} / 共 {size:>5})")
            elif mode == "UPTIME":
                uptime = format_uptime(line)
                mode = ""
                
        data_disk_str = "\n".join(data_disks) if data_disks else "无数据盘"
        data_list.append({
            "服务器IP": host, "系统版本": os_ver, "运行时间": uptime,
            "CPU状态": cpu, "内存状态": mem_str,
            "系统盘(/或C盘)": sys_disk, "数据盘": data_disk_str
        })
    return data_list

def get_windows_data():
    ps_cmd = "$os=(Get-WmiObject Win32_OperatingSystem).Caption.Replace('Microsoft Windows ','Win ').Replace(' 专业版',' Pro'); $cpu=(Get-WmiObject Win32_Processor | Measure-Object -Property LoadPercentage -Average).Average; $mem=Get-WmiObject Win32_OperatingSystem; $memTotal=[math]::Round($mem.TotalVisibleMemorySize/1024,0); $memUsed=$memTotal-[math]::Round($mem.FreePhysicalMemory/1024,0); $memRemain=$memTotal-$memUsed; if($memTotal -eq 0){$pct=0}else{$pct=[math]::Round(($memUsed/$memTotal)*100,1)}; $disks=Get-WmiObject Win32_LogicalDisk -Filter \"DriveType=3\"; $diskStr=\"\"; foreach($d in $disks){$size=[math]::Round($d.Size/1GB,0);$free=[math]::Round($d.FreeSpace/1GB,0);if($size -eq 0){$dpct=0}else{$dpct=[math]::Round((($size-$free)/$size)*100,0)};$diskStr+=\"$($d.DeviceID)|$($size)|$($free)|$dpct,\"}; $boot=$mem.LastBootUpTime; $now=Get-Date; $up=$now-[System.Management.ManagementDateTimeConverter]::ToDateTime($boot); $upSecs=[math]::Round($up.TotalSeconds,0); Write-Output \"---OS---`n$os`n---CPU---`n$cpu`n---MEM---`n$memTotal $memRemain $pct`n---DISK---`n$diskStr`n---UPTIME---`n$upSecs\""
    
    json_data = run_ansible_tree("windows", "win_shell", ps_cmd)
    data_list = []
    for host, info in json_data.items():
        if info.get('failed') or info.get('unreachable'): continue
        stdout = info.get('stdout', '')
        os_ver = cpu = mem_str = sys_disk = data_disk_str = uptime = "获取失败"
        data_disks = []
        mode = ""
        for line in stdout.split('\n'):
            line = line.strip()
            if not line: continue
            if line.startswith("---"): mode = line.replace("-", ""); continue

            if mode == "OS": os_ver = line; mode = ""
            elif mode == "CPU":
                try: cpu = f"{float(line):>5.1f}%"
                except: cpu = line
                mode = ""
            elif mode == "MEM":
                parts = line.split()
                if len(parts) >= 3:
                    total, remain, pct = float(parts[0]), float(parts[1]), float(parts[2])
                    mem_str = f"{pct:>4.1f}% (余 {remain/1024:>5.1f}G / 共 {total/1024:>5.1f}G)"
            elif mode == "DISK":
                for d in line.split(','):
                    if d:
                        parts = d.split('|')
                        if len(parts) == 4:
                            drive, size, free, pct = parts[0], parts[1], parts[2], parts[3]
                            format_str = f"{pct:>3}% (余 {str(free)+'G':>5} / 共 {str(size)+'G':>5})"
                            if drive == "C:": sys_disk = format_str
                            else: data_disks.append(f"{drive:<12} {format_str}")
            elif mode == "UPTIME":
                uptime = format_uptime(line)
                mode = ""
                
        data_disk_str = "\n".join(data_disks) if data_disks else "无数据盘"
        data_list.append({
            "服务器IP": host, "系统版本": os_ver, "运行时间": uptime,
            "CPU状态": cpu, "内存状态": mem_str,
            "系统盘(/或C盘)": sys_disk, "数据盘": data_disk_str
        })
    return data_list

def ip_to_int(ip_str):
    try: return struct.unpack("!I", socket.inet_aton(ip_str))[0]
    except: return 0

def sort_logic(item):
    ip = item.get("服务器IP", "")
    if ip == "10.0.0.168": return (1, ip_to_int(ip))
    elif ip == "10.0.0.199": return (2, ip_to_int(ip))
    else: return (0, ip_to_int(ip))

def beautify_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(bold=True, name="Microsoft YaHei", size=11)
    data_font = Font(name="Consolas", size=11)
    
    # 定义双阶梯高亮颜色
    warning_font = Font(name="Consolas", size=11, color="E69138", bold=True) # 琥珀橘黄 (80%~89%)
    alert_font = Font(name="Consolas", size=11, color="FF0000", bold=True)   # 爆红 (>=90%)
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    col_widths = {"A": 16, "B": 16, "C": 15, "D": 12, "E": 35, "F": 28, "G": 45}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 42
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = center_align
            cell.font = data_font
            
            val = str(cell.value)
            col_name = ws.cell(row=1, column=col).value
            
            try:
                # 内存双阶梯预警
                if col_name == "内存状态" and "%" in val:
                    pct = float(val.split('%')[0].strip())
                    if pct >= 90.0: cell.font = alert_font
                    elif pct >= 80.0: cell.font = warning_font
                    
                # 磁盘双阶梯预警 (多盘符提取最高使用率)
                if col_name in ["系统盘(/或C盘)", "数据盘"]:
                    max_pct = 0.0
                    for line in val.split('\n'):
                        if "%" in line:
                            pct_str = line.split('%')[0].split()[-1]
                            p = float(pct_str)
                            if p > max_pct: max_pct = p
                            
                    if max_pct >= 90.0: cell.font = alert_font
                    elif max_pct >= 80.0: cell.font = warning_font
            except: pass

    wb.save(file_path)

def main():
    print("🚀 开始收集全网数据...")
    all_data = get_linux_data() + get_windows_data()
    if not all_data: return

    all_data.sort(key=sort_logic)
    df = pd.DataFrame(all_data)
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    df.to_excel(EXCEL_PATH, index=False)
    beautify_excel(EXCEL_PATH)
    print(f"✅ Excel 报告已生成并完成双阶梯告警排版: {EXCEL_PATH}")

if __name__ == "__main__":
    main()